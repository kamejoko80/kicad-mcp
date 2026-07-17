"""Tests for BOM Mouser lookup helpers."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import scripts.build_bom_cost_excel_mouser as bom_cost


class BomMouserLookupTests(unittest.TestCase):
    def test_lookup_uses_mcp_only(self) -> None:
        with patch.object(bom_cost.asyncio, "run", side_effect=[True, {"Y": {"found": True}}]) as run_mock:
            result = bom_cost.lookup_mouser_prices([], bom_cost.DEFAULT_MCP_URL)
        self.assertEqual(run_mock.call_count, 2)
        self.assertEqual(result, {"Y": {"found": True}})

    def test_lookup_exits_when_mcp_unconfigured(self) -> None:
        with patch.object(bom_cost.asyncio, "run", return_value=False):
            with self.assertRaises(SystemExit):
                bom_cost.lookup_mouser_prices([], bom_cost.DEFAULT_MCP_URL)

    def test_manufacturer_variants_include_aliases(self) -> None:
        variants = bom_cost.manufacturer_variants("Murata Electronics", "TDK")
        self.assertIn("Murata", variants)
        self.assertIn("TDK Corporation", variants)

    def test_unit_price_for_order_qty_uses_mouser_break_tiers(self) -> None:
        breaks = [(1, 0.10), (10, 0.02), (100, 0.017), (2500, 0.003)]
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 1), 0.10)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 9), 0.10)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 10), 0.02)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 100), 0.017)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 5000), 0.003)

    def test_excel_tiered_unit_price_formula(self) -> None:
        formula = bom_cost.excel_tiered_unit_price_formula(
            "G6",
            [(1, 0.10), (100, 0.017)],
        )
        self.assertIn("INDEX", formula)
        self.assertIn("IFERROR", formula)
        self.assertIn("MATCH(G6", formula)
        self.assertIn("0.1", formula)
        self.assertIn("0.017", formula)

    def test_excel_formula_handles_order_qty_below_minimum_break(self) -> None:
        formula = bom_cost.excel_tiered_unit_price_formula("G6", [(2500, 0.05)])
        self.assertIn("IFERROR", formula)
        self.assertIn("2500", formula)

    def test_is_mouser_in_stock_uses_stock_quantity(self) -> None:
        self.assertTrue(
            bom_cost.is_mouser_in_stock({"stock_quantity": "421", "availability": "421 In Stock"})
        )
        self.assertFalse(
            bom_cost.is_mouser_in_stock(
                {"stock_quantity": "0", "availability": "Factory Lead Time"}
            )
        )

    def test_format_availability_label_normalizes_out_of_stock(self) -> None:
        label = bom_cost.format_availability_label(
            {"stock_quantity": "0", "availability": "Factory Lead Time"}
        )
        self.assertEqual(label, "Out of Stock")

    def test_normalized_price_breaks_skip_na_prices(self) -> None:
        from kicad_mcp.library.models import ComponentRecord, PriceBreak

        record = ComponentRecord(
            provider="mouser",
            distributor_part_number="TEST",
            manufacturer_part_number="TEST",
            manufacturer="Test",
            description="",
            price_breaks=[
                PriceBreak(quantity=1, price="NA"),
                PriceBreak(quantity=2500, price="$0.05"),
            ],
        )
        breaks = bom_cost.normalized_price_breaks(record)
        self.assertEqual(breaks, [(2500, 0.05)])

    def test_build_workbook_splits_instock_and_outstock_tables(self) -> None:
        bom_rows = [
            bom_cost.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "IN-STOCK-MPN", 1, ""),
            bom_cost.BomRow(2, "C2", "100nF", "C_0402", "YAGEO", "OUT-STOCK-MPN", 1, ""),
            bom_cost.BomRow(3, "C3", "100nF", "C_0402", "YAGEO", "MISSING-MPN", 1, ""),
        ]
        lookup = {
            "IN-STOCK-MPN": {
                "found": True,
                "price_breaks": [(1, 0.10)],
                "mouser_pn": "111",
                "stock_quantity": "100",
                "availability": "100 In Stock",
            },
            "OUT-STOCK-MPN": {
                "found": True,
                "price_breaks": [(1, 0.05)],
                "mouser_pn": "222",
                "stock_quantity": "0",
                "availability": "Factory Lead Time",
            },
            "MISSING-MPN": bom_cost.empty_lookup_entry(),
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost.xlsx"
            bom_cost.build_workbook(bom_rows, lookup, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            titles = [
                str(ws.cell(row, 1).value)
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
            ]
            self.assertIn("Placed Components — Available on Mouser (In Stock)", titles)
            self.assertIn("Placed Components — Available on Mouser (Out of Stock)", titles)
            self.assertNotIn("Placed Components — Available on Mouser", titles)
            self.assertIn("Placed Components — Not Available on Mouser (Unit Price = 0)", titles)

    def test_pick_exact_dict_does_not_fallback_to_first_fuzzy_match(self) -> None:
        records = [
            {"manufacturer_part_number": "RT4730WSC", "distributor_part_number": "C425170"},
        ]
        self.assertIsNone(bom_cost.pick_exact_dict(records, "AAT1556-WSC-T"))

    def test_unavailable_table_includes_bom_and_order_qty_columns(self) -> None:
        bom_rows = [
            bom_cost.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "MISSING-MPN", 2, ""),
        ]
        lookup = {"MISSING-MPN": bom_cost.empty_lookup_entry()}
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost.xlsx"
            bom_cost.build_workbook(bom_rows, lookup, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            header_row = next(
                row
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
                == "Placed Components — Not Available on Mouser (Unit Price = 0)"
            ) + 1
            self.assertEqual(ws.cell(header_row, 6).value, "BOM Qty / Board")
            self.assertEqual(ws.cell(header_row, 7).value, "Order Qty (BOM × PCBA)")
            data_row = header_row + 1
            self.assertEqual(ws.cell(data_row, 6).value, 2)
            self.assertEqual(ws.cell(data_row, 7).value, f"=F{data_row}*B3")


if __name__ == "__main__":
    unittest.main()
