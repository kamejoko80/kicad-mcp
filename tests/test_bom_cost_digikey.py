"""Tests for BOM DigiKey lookup helpers."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import scripts.build_bom_cost_excel_digikeys as bom_cost


class BomDigiKeyLookupTests(unittest.TestCase):
    def test_lookup_uses_mcp_only(self) -> None:
        with patch.object(bom_cost.asyncio, "run", side_effect=[True, {"Y": {"found": True}}]) as run_mock:
            result = bom_cost.lookup_digikey_prices([], bom_cost.DEFAULT_MCP_URL)
        self.assertEqual(run_mock.call_count, 2)
        self.assertEqual(result, {"Y": {"found": True}})

    def test_lookup_exits_when_mcp_unconfigured(self) -> None:
        with patch.object(bom_cost.asyncio, "run", return_value=False):
            with self.assertRaises(SystemExit):
                bom_cost.lookup_digikey_prices([], bom_cost.DEFAULT_MCP_URL)

    def test_unit_price_for_order_qty_uses_digikey_break_tiers(self) -> None:
        breaks = [(10000, 0.0059), (20000, 0.0053), (100000, 0.0045)]
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 9999), 0.0059)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 10000), 0.0059)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 20000), 0.0053)
        self.assertEqual(bom_cost.unit_price_for_order_qty(breaks, 100000), 0.0045)

    def test_is_digikey_in_stock_uses_stock_quantity(self) -> None:
        self.assertTrue(
            bom_cost.is_digikey_in_stock(
                {"stock_quantity": "2768005", "availability": "2768005 In Stock"}
            )
        )
        self.assertFalse(
            bom_cost.is_digikey_in_stock({"stock_quantity": "0", "availability": "Out of Stock"})
        )

    def test_format_availability_label_normalizes_out_of_stock(self) -> None:
        label = bom_cost.format_availability_label(
            {"stock_quantity": "0", "availability": "Backorder"}
        )
        self.assertEqual(label, "Out of Stock")

    def test_build_workbook_splits_instock_and_outstock_tables(self) -> None:
        bom_rows = [
            bom_cost.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "IN-STOCK-MPN", 1, ""),
            bom_cost.BomRow(2, "C2", "100nF", "C_0402", "YAGEO", "OUT-STOCK-MPN", 1, ""),
            bom_cost.BomRow(3, "C3", "100nF", "C_0402", "YAGEO", "MISSING-MPN", 1, ""),
        ]
        lookup = {
            "IN-STOCK-MPN": {
                "found": True,
                "price_breaks": [(1, 0.10)],
                "digikey_pn": "111",
                "stock_quantity": "100",
                "availability": "100 In Stock",
            },
            "OUT-STOCK-MPN": {
                "found": True,
                "price_breaks": [(1, 0.05)],
                "digikey_pn": "222",
                "stock_quantity": "0",
                "availability": "Backorder",
            },
            "MISSING-MPN": bom_cost.empty_lookup_entry(),
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost.xlsx"
            bom_cost.build_workbook(bom_rows, lookup, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            titles = [
                str(ws.cell(row, 1).value)
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
            ]
            self.assertIn("Placed Components — Available on DigiKey (In Stock)", titles)
            self.assertIn("Placed Components — Available on DigiKey (Out of Stock)", titles)
            self.assertNotIn("Placed Components — Available on DigiKey", titles)

    def test_unavailable_table_includes_bom_and_order_qty_columns(self) -> None:
        bom_rows = [
            bom_cost.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "MISSING-MPN", 2, ""),
        ]
        lookup = {"MISSING-MPN": bom_cost.empty_lookup_entry()}
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost.xlsx"
            bom_cost.build_workbook(bom_rows, lookup, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            header_row = next(
                row
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
                == "Placed Components — Not Available on DigiKey (Unit Price = 0)"
            ) + 1
            self.assertEqual(ws.cell(header_row, 6).value, "BOM Qty / Board")
            self.assertEqual(ws.cell(header_row, 7).value, "Order Qty (BOM × PCBA)")
            data_row = header_row + 1
            self.assertEqual(ws.cell(data_row, 6).value, 2)
            self.assertEqual(ws.cell(data_row, 7).value, f"=F{data_row}*B3")


if __name__ == "__main__":
    unittest.main()
