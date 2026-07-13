"""Tests for combined BOM cost cascading across DigiKey, Mouser, and LCSC."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import scripts.build_bom_cost_excel_all as bom_all


def _make_lookup(
    *,
    found: bool = True,
    in_stock: bool = True,
    pn_key: str = "digikey_pn",
    pn: str = "TEST-PN",
    price: float = 0.10,
) -> dict:
    availability = "100 In Stock" if in_stock else "Out of Stock"
    stock = "100" if in_stock else "0"
    entry: dict = {
        "found": found,
        "price_breaks": [(1, price)] if found else [],
        pn_key: pn,
        "stock_quantity": stock,
        "availability": availability,
    }
    if not found:
        entry = bom_all.empty_digikey_lookup()
    elif not in_stock:
        entry["found"] = True
        entry["price_breaks"] = [(1, price)]
        entry[pn_key] = pn
        entry["stock_quantity"] = "0"
        entry["availability"] = "Out of Stock"
    return entry


class ResolvePriorityTests(unittest.TestCase):
    def test_default_priorities(self) -> None:
        order = bom_all.resolve_priority_order(1, 2, 3)
        self.assertEqual(order, ["digikey", "mouser", "lcsc"])

    def test_custom_priorities(self) -> None:
        order = bom_all.resolve_priority_order(3, 1, 2)
        self.assertEqual(order, ["mouser", "lcsc", "digikey"])

    def test_invalid_priorities_raise(self) -> None:
        with self.assertRaises(SystemExit):
            bom_all.resolve_priority_order(1, 1, 2)


class CascadeAssignTests(unittest.TestCase):
    def _bom_rows(self) -> list[bom_all.BomRow]:
        return [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "DK-INSTOCK", 1, ""),
            bom_all.BomRow(2, "C2", "10uF", "C_0402", "YAGEO", "DK-OUTSTOCK", 1, ""),
            bom_all.BomRow(3, "C3", "1uH", "L_0402", "Murata", "MOUSER-ONLY", 1, ""),
            bom_all.BomRow(4, "C4", "0R", "R_0402", "YAGEO", "LCSC-ONLY", 1, ""),
            bom_all.BomRow(5, "C5", "10k", "R_0402", "YAGEO", "MISSING", 1, ""),
        ]

    def test_instock_on_prio1_stays_on_prio1(self) -> None:
        rows = self._bom_rows()
        lookups = {
            "digikey": {
                "DK-INSTOCK": _make_lookup(pn_key="digikey_pn", pn="DK-111", in_stock=True),
                "DK-OUTSTOCK": _make_lookup(pn_key="digikey_pn", pn="DK-222", in_stock=False),
            },
            "mouser": {
                "DK-OUTSTOCK": _make_lookup(pn_key="mouser_pn", pn="M-222", in_stock=True),
                "MOUSER-ONLY": _make_lookup(pn_key="mouser_pn", pn="M-333", in_stock=True),
            },
            "lcsc": {
                "LCSC-ONLY": _make_lookup(pn_key="lcsc_pn", pn="C-444", in_stock=True),
            },
        }
        result = bom_all.assign_parts_cascade(rows, ["digikey", "mouser", "lcsc"], lookups)

        dk_mpns = [row.mpn for row, _ in result.instock_by_provider["digikey"]]
        self.assertEqual(dk_mpns, ["DK-INSTOCK"])

        mouser_mpns = [row.mpn for row, _ in result.instock_by_provider["mouser"]]
        self.assertEqual(mouser_mpns, ["DK-OUTSTOCK", "MOUSER-ONLY"])

        lcsc_mpns = [row.mpn for row, _ in result.instock_by_provider["lcsc"]]
        self.assertEqual(lcsc_mpns, ["LCSC-ONLY"])

        unavailable_mpns = [row.mpn for row in result.unavailable]
        self.assertEqual(unavailable_mpns, ["MISSING"])

    def test_out_of_stock_on_prio1_cascades_to_prio2(self) -> None:
        rows = [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "PART-A", 1, ""),
        ]
        lookups = {
            "digikey": {
                "PART-A": _make_lookup(pn_key="digikey_pn", in_stock=False),
            },
            "mouser": {
                "PART-A": _make_lookup(pn_key="mouser_pn", pn="M-A", in_stock=True, price=0.05),
            },
        }
        result = bom_all.assign_parts_cascade(rows, ["digikey", "mouser", "lcsc"], lookups)

        self.assertNotIn("digikey", result.instock_by_provider)
        self.assertEqual(len(result.instock_by_provider["mouser"]), 1)
        self.assertEqual(result.instock_by_provider["mouser"][0][0].mpn, "PART-A")
        self.assertEqual(result.unavailable, [])

    def test_not_found_on_prio1_cascades_to_prio2(self) -> None:
        rows = [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "PART-B", 1, ""),
        ]
        lookups = {
            "digikey": {
                "PART-B": bom_all.empty_digikey_lookup(),
            },
            "mouser": {
                "PART-B": _make_lookup(pn_key="mouser_pn", pn="M-B", in_stock=True),
            },
        }
        result = bom_all.assign_parts_cascade(rows, ["digikey", "mouser", "lcsc"], lookups)

        self.assertNotIn("digikey", result.instock_by_provider)
        self.assertEqual(result.instock_by_provider["mouser"][0][0].mpn, "PART-B")

    def test_each_part_appears_in_only_one_table(self) -> None:
        rows = self._bom_rows()
        lookups = {
            "digikey": {
                "DK-INSTOCK": _make_lookup(pn_key="digikey_pn", in_stock=True),
                "DK-OUTSTOCK": _make_lookup(pn_key="digikey_pn", in_stock=False),
                "MOUSER-ONLY": _make_lookup(pn_key="digikey_pn", in_stock=False),
                "LCSC-ONLY": bom_all.empty_digikey_lookup(),
                "MISSING": bom_all.empty_digikey_lookup(),
            },
            "mouser": {
                "DK-OUTSTOCK": _make_lookup(pn_key="mouser_pn", in_stock=True),
                "MOUSER-ONLY": _make_lookup(pn_key="mouser_pn", in_stock=True),
                "LCSC-ONLY": bom_all.empty_mouser_lookup(),
                "MISSING": bom_all.empty_mouser_lookup(),
            },
            "lcsc": {
                "LCSC-ONLY": _make_lookup(pn_key="lcsc_pn", in_stock=True),
                "MISSING": bom_all.empty_lcsc_lookup(),
            },
        }
        result = bom_all.assign_parts_cascade(rows, ["digikey", "mouser", "lcsc"], lookups)

        seen_mpns: list[str] = []
        for items in result.instock_by_provider.values():
            for row, _ in items:
                seen_mpns.append(row.mpn)
        seen_mpns.extend(row.mpn for row in result.unavailable)
        self.assertEqual(len(seen_mpns), len(set(seen_mpns)))
        self.assertEqual(len(seen_mpns), 5)


class CascadeLookupTests(unittest.TestCase):
    def test_cascade_lookup_calls_providers_in_priority_order(self) -> None:
        rows = [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "PART-X", 1, ""),
        ]
        call_order: list[str] = []

        def fake_lookup(provider: str):
            def _fn(remaining_rows, mcp_url):
                call_order.append(provider)
                if provider == "digikey":
                    return {"PART-X": bom_all.empty_digikey_lookup()}
                if provider == "mouser":
                    return {
                        "PART-X": {
                            "found": True,
                            "price_breaks": [(1, 0.10)],
                            "mouser_pn": "M-X",
                            "stock_quantity": "50",
                            "availability": "50 In Stock",
                        }
                    }
                return {}

            return _fn

        with patch.dict(
            bom_all.LOOKUP_FUNCTIONS,
            {
                "digikey": fake_lookup("digikey"),
                "mouser": fake_lookup("mouser"),
                "lcsc": fake_lookup("lcsc"),
            },
        ):
            result = bom_all.cascade_lookup_prices(rows, ["digikey", "mouser", "lcsc"], "http://test/mcp")

        self.assertEqual(call_order, ["digikey", "mouser"])
        self.assertNotIn("lcsc", call_order)
        self.assertEqual(result.instock_by_provider["mouser"][0][0].mpn, "PART-X")


class BuildWorkbookTests(unittest.TestCase):
    def test_workbook_table_titles_present(self) -> None:
        rows = [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "DK-INSTOCK", 1, ""),
            bom_all.BomRow(2, "C2", "10uF", "C_0402", "YAGEO", "MOUSER-ONLY", 1, ""),
            bom_all.BomRow(3, "C3", "1uH", "L_0402", "Murata", "LCSC-ONLY", 1, ""),
            bom_all.BomRow(4, "C4", "0R", "R_0402", "YAGEO", "MISSING", 1, ""),
            bom_all.BomRow(5, "TP1", "TEST", "TP", "YAGEO", "DNP-MPN", 1, "Do not place"),
        ]
        lookups = {
            "digikey": {
                "DK-INSTOCK": _make_lookup(pn_key="digikey_pn", in_stock=True),
                "MOUSER-ONLY": bom_all.empty_digikey_lookup(),
                "LCSC-ONLY": bom_all.empty_digikey_lookup(),
                "MISSING": bom_all.empty_digikey_lookup(),
            },
            "mouser": {
                "MOUSER-ONLY": _make_lookup(pn_key="mouser_pn", in_stock=True),
                "LCSC-ONLY": bom_all.empty_mouser_lookup(),
                "MISSING": bom_all.empty_mouser_lookup(),
            },
            "lcsc": {
                "LCSC-ONLY": _make_lookup(pn_key="lcsc_pn", in_stock=True),
                "MISSING": bom_all.empty_lcsc_lookup(),
            },
        }
        cascade = bom_all.assign_parts_cascade(rows, ["digikey", "mouser", "lcsc"], lookups)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost_all.xlsx"
            bom_all.build_workbook(rows, cascade, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            titles = [
                str(ws.cell(row, 1).value)
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
            ]
            self.assertIn("Placed Components — Available on DigiKey (In Stock)", titles)
            self.assertIn("Placed Components — Available on Mouser (In Stock)", titles)
            self.assertIn("Placed Components — Available on LCSC (In Stock)", titles)
            self.assertIn(
                "Placed Components — Not Available on above distributors (Unit Price = 0)",
                titles,
            )
            self.assertIn("Combined BOM Cost Summary", titles)
            self.assertIn("Do Not Place (DNP)", titles)
            self.assertNotIn("Placed Components — Available on DigiKey (Out of Stock)", titles)
            self.assertNotIn("Placed Components — Available on Mouser (Out of Stock)", titles)

    def test_workbook_respects_custom_priority_order(self) -> None:
        rows = [
            bom_all.BomRow(1, "C1", "100nF", "C_0402", "YAGEO", "PART-Z", 1, ""),
        ]
        lookups = {
            "lcsc": {
                "PART-Z": _make_lookup(pn_key="lcsc_pn", pn="C-Z", in_stock=True),
            },
            "digikey": {
                "PART-Z": bom_all.empty_digikey_lookup(),
            },
            "mouser": {
                "PART-Z": bom_all.empty_mouser_lookup(),
            },
        }
        cascade = bom_all.assign_parts_cascade(rows, ["lcsc", "digikey", "mouser"], lookups)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bom_cost_all.xlsx"
            bom_all.build_workbook(rows, cascade, output_path, "test.csv")
            from openpyxl import load_workbook

            ws = load_workbook(output_path)["BOM Cost"]
            titles = [
                str(ws.cell(row, 1).value)
                for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value
            ]
            lcsc_idx = titles.index("Placed Components — Available on LCSC (In Stock)")
            digikey_present = "Placed Components — Available on DigiKey (In Stock)" in titles
            self.assertLess(lcsc_idx, titles.index("Combined BOM Cost Summary"))
            self.assertFalse(digikey_present)


if __name__ == "__main__":
    unittest.main()
