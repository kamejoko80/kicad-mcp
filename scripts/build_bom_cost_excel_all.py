"""Build a combined BOM cost Excel workbook with priority-based distributor cascading.

Usage:
  uv run --with openpyxl python scripts/build_bom_cost_excel_all.py <bom.csv> [output.xlsx]

Searches DigiKey, Mouser, and LCSC via the KiCad MCP server (default: http://127.0.0.1:8500/mcp).
Parts cascade to the next-priority distributor when not found or out-of-stock on the current one.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Allow running from repo root without installing the package.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.build_bom_cost_excel_digikeys import (
    DEFAULT_MCP_URL,
    BomRow,
    empty_lookup_entry as empty_digikey_lookup,
    group_placed_by_mpn,
    is_digikey_in_stock,
    lookup_digikey_prices,
    parse_bom_csv,
    style_header,
    write_digikey_placed_table,
    write_table_title,
)
from scripts.build_bom_cost_excel_lcsc import (
    empty_lookup_entry as empty_lcsc_lookup,
    is_lcsc_in_stock,
    lookup_lcsc_prices,
    write_lcsc_placed_table,
)
from scripts.build_bom_cost_excel_mouser import (
    empty_lookup_entry as empty_mouser_lookup,
    is_mouser_in_stock,
    lookup_mouser_prices,
    write_mouser_placed_table,
)

DEFAULT_DIGIKEYS_PRIO = 1
DEFAULT_MOUSER_PRIO = 2
DEFAULT_LCSC_PRIO = 3

PROVIDER_LABELS: dict[str, str] = {
    "digikey": "DigiKey",
    "mouser": "Mouser",
    "lcsc": "LCSC",
}

IN_STOCK_TABLE_TITLES: dict[str, str] = {
    "digikey": "Placed Components — Available on DigiKey (In Stock)",
    "mouser": "Placed Components — Available on Mouser (In Stock)",
    "lcsc": "Placed Components — Available on LCSC (In Stock)",
}

INSTOCK_TOTAL_LABELS: dict[str, str] = {
    "digikey": "Total DigiKey (instock)",
    "mouser": "Total Mouser (instock)",
    "lcsc": "Total LCSC (instock)",
}

WRITE_PLACED_TABLE = {
    "digikey": write_digikey_placed_table,
    "mouser": write_mouser_placed_table,
    "lcsc": write_lcsc_placed_table,
}

LOOKUP_FUNCTIONS = {
    "digikey": lookup_digikey_prices,
    "mouser": lookup_mouser_prices,
    "lcsc": lookup_lcsc_prices,
}

IS_IN_STOCK = {
    "digikey": is_digikey_in_stock,
    "mouser": is_mouser_in_stock,
    "lcsc": is_lcsc_in_stock,
}

EMPTY_LOOKUP = {
    "digikey": empty_digikey_lookup,
    "mouser": empty_mouser_lookup,
    "lcsc": empty_lcsc_lookup,
}


@dataclass
class CascadeResult:
    instock_by_provider: dict[str, list[tuple[BomRow, dict[str, object]]]] = field(
        default_factory=dict
    )
    unavailable: list[BomRow] = field(default_factory=list)
    priority_order: list[str] = field(default_factory=list)


def resolve_priority_order(
    digikeys_prio: int,
    mouser_prio: int,
    lcsc_prio: int,
) -> list[str]:
    """Return provider names sorted by priority (1 = first)."""
    priorities = {
        "digikey": digikeys_prio,
        "mouser": mouser_prio,
        "lcsc": lcsc_prio,
    }
    values = list(priorities.values())
    if sorted(values) != [1, 2, 3]:
        raise SystemExit(
            "Distributor priorities must be a permutation of 1, 2, and 3 "
            f"(got digikey={digikeys_prio}, mouser={mouser_prio}, lcsc={lcsc_prio})."
        )
    return [name for name, _prio in sorted(priorities.items(), key=lambda item: item[1])]


def assign_parts_cascade(
    bom_rows: list[BomRow],
    priority_order: list[str],
    lookups: dict[str, dict[str, dict[str, object]]],
) -> CascadeResult:
    """Assign each placed part to the first in-stock distributor in priority order."""
    placed = [row for row in bom_rows if not row.is_dnp]
    remaining = placed[:]
    instock_by_provider: dict[str, list[tuple[BomRow, dict[str, object]]]] = {}

    for provider in priority_order:
        if not remaining:
            break
        is_in_stock = IS_IN_STOCK[provider]
        empty_fn = EMPTY_LOOKUP[provider]
        lookup = lookups.get(provider, {})
        instock_items: list[tuple[BomRow, dict[str, object]]] = []
        still_remaining: list[BomRow] = []

        for row in remaining:
            info = lookup.get(row.mpn, empty_fn())
            if info.get("found") and is_in_stock(info):
                instock_items.append((row, info))
            else:
                still_remaining.append(row)

        if instock_items:
            instock_by_provider[provider] = instock_items
        remaining = still_remaining

    return CascadeResult(
        instock_by_provider=instock_by_provider,
        unavailable=remaining,
        priority_order=priority_order,
    )


def _empty_lookup_for_rows(rows: list[BomRow], empty_fn) -> dict[str, dict[str, object]]:
    grouped = group_placed_by_mpn(rows)
    return {mpn: empty_fn() for mpn, _ in grouped}


def safe_lookup_provider(
    provider: str,
    rows: list[BomRow],
    mcp_url: str,
) -> dict[str, dict[str, object]]:
    """Look up prices for rows on one distributor; return empty entries if MCP unavailable."""
    lookup_fn = LOOKUP_FUNCTIONS[provider]
    empty_fn = EMPTY_LOOKUP[provider]
    label = PROVIDER_LABELS[provider]
    try:
        return lookup_fn(rows, mcp_url)
    except SystemExit:
        print(f"Warning: {label} is not available via MCP; cascading remaining parts to next distributor.")
        return _empty_lookup_for_rows(rows, empty_fn)


def cascade_lookup_prices(
    bom_rows: list[BomRow],
    priority_order: list[str],
    mcp_url: str,
) -> CascadeResult:
    """Search distributors in priority order; cascade unmatched/out-of-stock parts."""
    placed = [row for row in bom_rows if not row.is_dnp]
    remaining = placed[:]
    lookups: dict[str, dict[str, dict[str, object]]] = {}

    for provider in priority_order:
        if not remaining:
            break
        label = PROVIDER_LABELS[provider]
        print(f"Searching {label} for {len(remaining)} remaining placed line(s)...")
        lookup = safe_lookup_provider(provider, remaining, mcp_url)
        lookups[provider] = lookup

        is_in_stock = IS_IN_STOCK[provider]
        empty_fn = EMPTY_LOOKUP[provider]
        still_remaining: list[BomRow] = []
        instock_count = 0
        for row in remaining:
            info = lookup.get(row.mpn, empty_fn())
            if info.get("found") and is_in_stock(info):
                instock_count += 1
            else:
                still_remaining.append(row)
        print(f"  {label}: {instock_count} in-stock, {len(still_remaining)} cascading to next distributor")
        remaining = still_remaining

    return assign_parts_cascade(bom_rows, priority_order, lookups)


def write_unavailable_table(
    ws,
    *,
    current_row: int,
    rows: list[BomRow],
    pcba_cell: str,
) -> tuple[int, int | None]:
    """Write the not-available table; return (next_row, total_row)."""
    write_table_title(
        ws,
        current_row,
        "Placed Components — Not Available on above distributors (Unit Price = 0)",
    )
    current_row += 1
    unavailable_headers = [
        "Line",
        "References",
        "Value",
        "Manufacturer",
        "MPN",
        "BOM Qty",
        "Unit Price (USD)",
        "Line Total / Board (USD)",
    ]
    for col, title in enumerate(unavailable_headers, start=1):
        style_header(ws.cell(row=current_row, column=col, value=title))
    current_row += 1
    unavailable_first_data_row = current_row

    for row in rows:
        ws.cell(row=current_row, column=1, value=row.line)
        ws.cell(row=current_row, column=2, value=row.references)
        ws.cell(row=current_row, column=3, value=row.value)
        ws.cell(row=current_row, column=4, value=row.manufacturer)
        ws.cell(row=current_row, column=5, value=row.mpn)
        ws.cell(row=current_row, column=6, value=row.quantity)
        zero_cell = ws.cell(row=current_row, column=7, value=0)
        zero_cell.number_format = "$0.0000"
        line_total = ws.cell(row=current_row, column=8, value=f"=F{current_row}*G{current_row}")
        line_total.number_format = "$0.0000"
        current_row += 1

    unavailable_last_data_row = current_row - 1
    total_row: int | None = None
    if rows:
        ws.cell(row=current_row, column=5, value="Total (Not available × PCBA Qty)").font = Font(bold=True)
        total_cell = ws.cell(
            row=current_row,
            column=8,
            value=f"=SUM(H{unavailable_first_data_row}:H{unavailable_last_data_row})*{pcba_cell}",
        )
        total_cell.font = Font(bold=True)
        total_cell.number_format = "$0.00"
        total_row = current_row
        current_row += 2
    else:
        current_row += 1

    return current_row, total_row


def build_workbook(
    bom_rows: list[BomRow],
    cascade: CascadeResult,
    output_path: Path,
    source_name: str,
) -> None:
    dnp_rows = [row for row in bom_rows if row.is_dnp]

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Cost"

    ws["A1"] = "BOM Cost Estimate (DigiKey / Mouser / LCSC)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Source BOM"
    ws["B2"] = source_name
    ws["A3"] = "PCBA Quantity"
    ws["B3"] = 1
    ws["B3"].number_format = "0"
    ws["A3"].font = Font(bold=True)
    ws["B3"].fill = PatternFill("solid", fgColor="FFF2CC")

    pcba_cell = "B3"
    current_row = 5
    instock_total_rows: dict[str, int] = {}

    for provider in cascade.priority_order:
        items = cascade.instock_by_provider.get(provider, [])
        if not items:
            continue
        write_fn = WRITE_PLACED_TABLE[provider]
        current_row, total_row = write_fn(
            ws,
            current_row=current_row,
            title=IN_STOCK_TABLE_TITLES[provider],
            items=items,
            pcba_cell=pcba_cell,
            total_label=INSTOCK_TOTAL_LABELS[provider],
        )
        if total_row is not None:
            instock_total_rows[provider] = total_row

    current_row, unavailable_total_row = write_unavailable_table(
        ws,
        current_row=current_row,
        rows=cascade.unavailable,
        pcba_cell=pcba_cell,
    )

    write_table_title(ws, current_row, "Combined BOM Cost Summary")
    current_row += 1
    summary_rows: list[int] = []

    for provider in cascade.priority_order:
        total_row = instock_total_rows.get(provider)
        if total_row is None:
            continue
        label = PROVIDER_LABELS[provider]
        ws.cell(row=current_row, column=5, value=f"Total — {label} (instock)").font = Font(bold=True)
        ws.cell(row=current_row, column=9, value=f"=I{total_row}").number_format = "$0.00"
        summary_rows.append(current_row)
        current_row += 1

    unavailable_total_ref = f"H{unavailable_total_row}" if unavailable_total_row else "0"
    ws.cell(row=current_row, column=5, value="Total — Not available (× PCBA Qty)").font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=f"={unavailable_total_ref}").number_format = "$0.00"
    unavailable_summary_row = current_row
    current_row += 1

    combined_parts = [f"I{row}" for row in summary_rows]
    combined_parts.append(f"I{unavailable_summary_row}")
    combined_formula = "=" + "+".join(combined_parts) if combined_parts else "=0"

    ws.cell(row=current_row, column=5, value="Combined Total (All PCBA)").font = Font(bold=True)
    combined_total_cell = ws.cell(row=current_row, column=9, value=combined_formula)
    combined_total_cell.font = Font(bold=True, size=12)
    combined_total_cell.number_format = "$0.00"
    combined_total_cell.fill = PatternFill("solid", fgColor="DDEBF7")
    combined_total_row = current_row
    current_row += 1

    ws.cell(row=current_row, column=5, value="Grand Total").font = Font(bold=True)
    grand_total_cell = ws.cell(row=current_row, column=9, value=f"=I{combined_total_row}")
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.number_format = "$0.00"
    grand_total_cell.fill = PatternFill("solid", fgColor="E2EFDA")
    current_row += 3

    write_table_title(ws, current_row, "Do Not Place (DNP)")
    current_row += 1
    dnp_headers = ["Line", "References", "Value", "Manufacturer", "MPN", "BOM Qty", "Note"]
    for col, title in enumerate(dnp_headers, start=1):
        style_header(ws.cell(row=current_row, column=col, value=title))
    current_row += 1
    for row in dnp_rows:
        ws.cell(row=current_row, column=1, value=row.line)
        ws.cell(row=current_row, column=2, value=row.references)
        ws.cell(row=current_row, column=3, value=row.value)
        ws.cell(row=current_row, column=4, value=row.manufacturer)
        ws.cell(row=current_row, column=5, value=row.mpn)
        ws.cell(row=current_row, column=6, value=row.quantity)
        ws.cell(row=current_row, column=7, value=row.note or "Do not place")
        current_row += 1

    widths = [8, 28, 18, 22, 28, 12, 16, 18, 16, 18, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A6"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_updated{output_path.suffix}")
        wb.save(fallback)
        print(f"Could not overwrite locked file. Saved to {fallback}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build combined BOM cost Excel with priority-based distributor cascading."
    )
    parser.add_argument("bom_csv", type=Path, help="Path to KiCad BOM CSV export")
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        type=Path,
        help="Output Excel path (default: <bom>_bom_cost_all.xlsx beside CSV)",
    )
    parser.add_argument(
        "--mcp-url",
        default=DEFAULT_MCP_URL,
        help=f"KiCad MCP server URL (default: {DEFAULT_MCP_URL})",
    )
    parser.add_argument(
        "--digikeys-prio",
        type=int,
        default=DEFAULT_DIGIKEYS_PRIO,
        help=f"DigiKey search priority (default: {DEFAULT_DIGIKEYS_PRIO})",
    )
    parser.add_argument(
        "--mouser-prio",
        type=int,
        default=DEFAULT_MOUSER_PRIO,
        help=f"Mouser search priority (default: {DEFAULT_MOUSER_PRIO})",
    )
    parser.add_argument(
        "--lcsc-prio",
        type=int,
        default=DEFAULT_LCSC_PRIO,
        help=f"LCSC search priority (default: {DEFAULT_LCSC_PRIO})",
    )
    args = parser.parse_args()

    bom_path = args.bom_csv.expanduser().resolve()
    if not bom_path.is_file():
        print(f"BOM file not found: {bom_path}", file=sys.stderr)
        return 1

    output_path = (
        args.output_xlsx.expanduser().resolve()
        if args.output_xlsx
        else bom_path.with_name(f"{bom_path.stem}_bom_cost_all.xlsx")
    )

    priority_order = resolve_priority_order(
        args.digikeys_prio,
        args.mouser_prio,
        args.lcsc_prio,
    )
    print(
        "Distributor priority: "
        + ", ".join(f"{PROVIDER_LABELS[p]}={i + 1}" for i, p in enumerate(priority_order))
    )

    bom_rows = parse_bom_csv(bom_path)
    placed_rows = [row for row in bom_rows if not row.is_dnp]
    print(f"Parsed {len(bom_rows)} BOM lines ({len(placed_rows)} placed, {len(bom_rows) - len(placed_rows)} DNP)")

    cascade = cascade_lookup_prices(bom_rows, priority_order, args.mcp_url.strip())
    build_workbook(bom_rows, cascade, output_path, bom_path.name)

    assigned = sum(len(items) for items in cascade.instock_by_provider.values())
    print(f"Wrote {output_path}")
    print(
        f"Assigned: {assigned} in-stock across distributors, "
        f"{len(cascade.unavailable)} not available"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
