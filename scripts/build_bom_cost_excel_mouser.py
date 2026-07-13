"""Build a BOM cost Excel workbook with Mouser pricing.

Usage:
  uv run --with openpyxl python scripts/build_bom_cost_excel_mouser.py <bom.csv> [output.xlsx]

Mouser part data is fetched only through the KiCad MCP server (default: http://127.0.0.1:8500/mcp).
Start the MCP server with MOUSER_API_KEY configured in that process before running this script.
This script never reads MOUSER_API_KEY from its own terminal environment.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Allow running from repo root without installing the package.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from kicad_mcp.library.providers.mouser import MAX_PART_NUMBERS

DEFAULT_MCP_URL = "http://127.0.0.1:8500/mcp"


@dataclass
class BomRow:
    line: int
    references: str
    value: str
    footprint: str
    manufacturer: str
    mpn: str
    quantity: int
    note: str

    @property
    def is_dnp(self) -> bool:
        if self.note.strip().casefold() == "do not place":
            return True
        return "(dnp)" in self.value.casefold()


def parse_bom_csv(path: Path) -> list[BomRow]:
    rows: list[BomRow] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for index, raw in enumerate(reader, start=1):
            mpn = (raw.get("Manufacturer_Part_Number") or "").strip()
            if not mpn:
                continue
            qty_raw = (raw.get("Quantity") or "0").strip()
            try:
                quantity = int(float(qty_raw))
            except ValueError:
                quantity = 0
            rows.append(
                BomRow(
                    line=index,
                    references=(raw.get("References") or "").strip(),
                    value=(raw.get("Value") or "").strip(),
                    footprint=(raw.get("Footprint") or "").strip(),
                    manufacturer=(raw.get("Manufacturer_Name") or "").strip(),
                    mpn=mpn,
                    quantity=quantity,
                    note=(raw.get("Note") or "").strip(),
                )
            )
    return rows


def parse_unit_price(price_text: str) -> float | None:
    cleaned = price_text.strip().replace("$", "").replace(",", "")
    if not cleaned or cleaned.casefold() in {"na", "n/a", "quote", "rfq", "-"}:
        return None
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if value < 0:
        return None
    return value


def normalized_price_breaks(record) -> list[tuple[int, float]]:
    breaks: list[tuple[int, float]] = []
    for item in record.price_breaks:
        try:
            quantity = int(item.quantity or 0)
        except (TypeError, ValueError):
            quantity = 0
        price = parse_unit_price(str(item.price))
        if quantity <= 0 or price is None:
            continue
        breaks.append((quantity, price))
    breaks.sort(key=lambda pair: pair[0])
    if not breaks:
        return breaks
    deduped: list[tuple[int, float]] = []
    for quantity, price in breaks:
        if deduped and deduped[-1][0] == quantity:
            deduped[-1] = (quantity, price)
        else:
            deduped.append((quantity, price))
    return deduped


def normalized_price_breaks_from_dict(item: dict) -> list[tuple[int, float]]:
    breaks: list[tuple[int, float]] = []
    for entry in item.get("price_breaks") or []:
        try:
            quantity = int(entry.get("quantity") or 0)
        except (TypeError, ValueError):
            quantity = 0
        price = parse_unit_price(str(entry.get("price", "")))
        if quantity <= 0 or price is None:
            continue
        breaks.append((quantity, price))
    breaks.sort(key=lambda pair: pair[0])
    return breaks


def usable_price_breaks(breaks: list[tuple[int, float]]) -> list[tuple[int, float]]:
    return [(quantity, price) for quantity, price in breaks if price > 0]


def unit_price_for_order_qty(breaks: list[tuple[int, float]], order_qty: int) -> float:
    if not breaks or order_qty <= 0:
        return 0.0
    unit_price = breaks[0][1]
    for break_qty, break_price in breaks:
        if break_qty <= order_qty:
            unit_price = break_price
        else:
            break
    return unit_price


def excel_tiered_unit_price_formula(order_qty_expr: str, breaks: list[tuple[int, float]]) -> str:
    usable = usable_price_breaks(breaks)
    if not usable:
        return "0"
    qty_values = ",".join(str(qty) for qty, _ in usable)
    price_values = ",".join(str(price) for _, price in usable)
    # IFERROR: order qty below the lowest Mouser break otherwise yields MATCH #N/A.
    return (
        f"=IF({order_qty_expr}<=0,0,"
        f"IFERROR(INDEX({{{price_values}}},MATCH({order_qty_expr},{{{qty_values}}},1)),"
        f"INDEX({{{price_values}}},1)))"
    )


def record_to_lookup_entry(record) -> dict[str, object]:
    breaks = normalized_price_breaks(record)
    return {
        "found": True,
        "price_breaks": breaks,
        "unit_price": unit_price_for_order_qty(breaks, 1),
        "mouser_pn": record.distributor_part_number,
        "availability": record.availability,
        "stock_quantity": record.stock_quantity,
        "description": record.description,
    }


def empty_lookup_entry() -> dict[str, object]:
    return {
        "found": False,
        "unit_price": 0.0,
        "mouser_pn": "",
        "availability": "",
        "stock_quantity": "",
        "description": "",
    }


def is_mouser_in_stock(info: dict[str, object]) -> bool:
    stock_text = str(info.get("stock_quantity", "")).strip()
    if stock_text:
        try:
            return int(float(stock_text)) > 0
        except ValueError:
            pass
    availability = str(info.get("availability", "")).casefold()
    if "out of stock" in availability:
        return False
    return "in stock" in availability


def format_availability_label(info: dict[str, object]) -> str:
    if is_mouser_in_stock(info):
        stock_text = str(info.get("stock_quantity", "")).strip()
        if stock_text:
            try:
                quantity = int(float(stock_text))
                if quantity > 0:
                    return f"{quantity} In Stock"
            except ValueError:
                pass
        availability = str(info.get("availability", "")).strip()
        if availability and "in stock" in availability.casefold():
            return availability
        return "In Stock"
    return "Out of Stock"


def pick_exact_dict(records: list[dict], mpn: str) -> dict | None:
    target = mpn.casefold()
    for item in records:
        if str(item.get("manufacturer_part_number", "")).casefold() == target:
            return item
    return records[0] if records else None


MANUFACTURER_ALIASES: dict[str, list[str]] = {
    "TDK": ["TDK Corporation", "TDK Electronics"],
    "Hirose": ["Hirose Electric", "Hirose Electric Co Ltd"],
    "Johanson Technology Inc": ["Johanson Technology"],
    "GLF Integrated Power": ["GLF Power", "GLF Power Inc"],
    "Advanced Analog Technology, Inc": ["Advanced Analog Technology", "Advanced Analog"],
    "Murata Electronics": ["Murata"],
    "TE Connectivity": ["TE Connectivity / Holsworthy", "TE"],
    "Sonny": ["Sony"],
}


def manufacturer_variants(*names: str) -> list[str]:
    variants: list[str] = []
    for name in names:
        cleaned = name.strip()
        if not cleaned:
            continue
        candidates = [
            cleaned,
            cleaned.split("/")[0].strip(),
            cleaned.split("/")[-1].strip() if "/" in cleaned else "",
            cleaned.split(",")[0].strip(),
            cleaned.replace(" Inc.", "").replace(" Inc", "").replace(" Corporation", "").strip(),
            cleaned.replace(" Electronics", "").strip(),
            cleaned.replace(", Inc", "").strip(),
        ]
        for candidate in candidates:
            if candidate and candidate not in variants:
                variants.append(candidate)
            for alias in MANUFACTURER_ALIASES.get(candidate, []):
                if alias not in variants:
                    variants.append(alias)
    return variants


def group_placed_by_mpn(rows: list[BomRow]) -> list[tuple[str, list[str]]]:
    order: list[str] = []
    manufacturers_by_mpn: dict[str, list[str]] = {}
    for row in rows:
        if row.is_dnp:
            continue
        if row.mpn not in manufacturers_by_mpn:
            order.append(row.mpn)
            manufacturers_by_mpn[row.mpn] = []
        if row.manufacturer and row.manufacturer not in manufacturers_by_mpn[row.mpn]:
            manufacturers_by_mpn[row.mpn].append(row.manufacturer)
    return [(mpn, manufacturers_by_mpn[mpn]) for mpn in order]


def apply_batch_results(
    lookup: dict[str, dict[str, object]],
    mpns: list[str],
    records: list,
) -> None:
    by_mpn: dict[str, object] = {}
    for record in records:
        key = record.manufacturer_part_number.casefold()
        if key not in by_mpn:
            by_mpn[key] = record
    for mpn in mpns:
        if lookup.get(mpn, {}).get("found"):
            continue
        record = by_mpn.get(mpn.casefold())
        if record:
            lookup[mpn] = record_to_lookup_entry(record)


async def mcp_search_part(
    session,
    *,
    mpn: str,
    manufacturers: list[str],
    tool_json,
) -> dict | None:
    attempts: list[dict[str, str]] = [{"part_number": mpn, "provider": "mouser", "match_mode": "Exact"}]
    for manufacturer_name in manufacturer_variants(*manufacturers):
        attempts.append(
            {
                "part_number": mpn,
                "provider": "mouser",
                "manufacturer": manufacturer_name,
                "match_mode": "Exact",
            }
        )

    for arguments in attempts:
        for retry in range(3):
            result = await session.call_tool("search_components_by_part_number", arguments)
            payload = tool_json(result)
            if payload.get("error"):
                if retry == 2:
                    print(f"  Warning for {mpn}: {payload['error']}")
                await asyncio.sleep(0.4 * (retry + 1))
                continue
            records = payload.get("records") or []
            match = pick_exact_dict(records, mpn)
            if match:
                return match
            errors = payload.get("errors") or []
            if errors and retry == 2:
                print(f"  Warning for {mpn}: {errors}")
            if errors and any("NotFound" in error for error in errors):
                break
            await asyncio.sleep(0.25)

    result = await session.call_tool(
        "search_components_by_keyword",
        {"keyword": mpn, "provider": "mouser", "records": 10},
    )
    payload = tool_json(result)
    if not payload.get("error"):
        match = pick_exact_dict(payload.get("records") or [], mpn)
        if match:
            return match
    return None


async def probe_mcp_mouser(mcp_url: str) -> bool:
    from mcp.client.session import ClientSession
    from mcp.client.streamable_http import streamable_http_client
    from mcp.types import CallToolResult

    def tool_json(result: CallToolResult) -> dict:
        text_parts: list[str] = []
        for block in result.content:
            text = getattr(block, "text", None)
            if text:
                text_parts.append(text)
        payload = json.loads("\n".join(text_parts))
        return payload if isinstance(payload, dict) else {}

    try:
        async with streamable_http_client(mcp_url) as (read_stream, write_stream, _get_session_id):
            async with ClientSession(read_stream, write_stream) as session:
                await session.initialize()
                result = await session.call_tool(
                    "get_component_provider_status",
                    {"provider": "mouser"},
                )
                payload = tool_json(result)
                providers = payload.get("providers") or []
                if providers:
                    return bool(providers[0].get("configured"))
                provider = payload.get("provider") or {}
                return bool(provider.get("configured"))
    except Exception as exc:
        print(f"Could not reach MCP server at {mcp_url}: {exc}")
        return False


async def mcp_call_part_search(session, arguments: dict[str, str], tool_json) -> dict:
    result = await session.call_tool("search_components_by_part_number", arguments)
    return tool_json(result)


async def lookup_mouser_prices_mcp(
    rows: list[BomRow],
    mcp_url: str,
) -> dict[str, dict[str, object]]:
    from mcp.client.session import ClientSession
    from mcp.client.streamable_http import streamable_http_client
    from mcp.types import CallToolResult

    def tool_json(result: CallToolResult) -> dict:
        text_parts: list[str] = []
        for block in result.content:
            text = getattr(block, "text", None)
            if text:
                text_parts.append(text)
        payload = json.loads("\n".join(text_parts))
        if not isinstance(payload, dict):
            raise RuntimeError("MCP tool returned unexpected payload.")
        return payload

    lookup: dict[str, dict[str, object]] = {}
    grouped = group_placed_by_mpn(rows)
    mpns = [mpn for mpn, _ in grouped]
    manufacturers_by_mpn = {mpn: manufacturers for mpn, manufacturers in grouped}

    async with streamable_http_client(mcp_url) as (read_stream, write_stream, _get_session_id):
        async with ClientSession(read_stream, write_stream) as session:
            await session.initialize()

            for start in range(0, len(mpns), MAX_PART_NUMBERS):
                batch = mpns[start : start + MAX_PART_NUMBERS]
                query = "|".join(batch)
                print(f"MCP batch query ({start + 1}-{start + len(batch)} of {len(mpns)}): {len(batch)} parts")
                payload = await mcp_call_part_search(
                    session,
                    {"part_number": query, "provider": "mouser", "match_mode": "Exact"},
                    tool_json,
                )
                if payload.get("error"):
                    print(f"  Warning for batch starting {batch[0]}: {payload['error']}")
                else:
                    records = [_dict_to_record(item) for item in payload.get("records") or []]
                    apply_batch_results(lookup, batch, records)
                await asyncio.sleep(1.0)

            missing = [mpn for mpn in mpns if not lookup.get(mpn, {}).get("found")]
            for index, mpn in enumerate(missing, start=1):
                print(f"MCP retry individually ({index}/{len(missing)}): {mpn}")
                item = await mcp_search_part(
                    session,
                    mpn=mpn,
                    manufacturers=manufacturers_by_mpn[mpn],
                    tool_json=tool_json,
                )
                if item:
                    record = _dict_to_record(item)
                    lookup[mpn] = record_to_lookup_entry(record)
                else:
                    lookup[mpn] = empty_lookup_entry()
                await asyncio.sleep(0.5)

    return lookup


def _dict_to_record(item: dict):
    from kicad_mcp.library.models import ComponentRecord, PriceBreak

    price_breaks = [
        PriceBreak(
            quantity=int(pb.get("quantity") or 0),
            price=str(pb.get("price", "")),
            currency=str(pb.get("currency", "USD")),
        )
        for pb in item.get("price_breaks") or []
    ]
    return ComponentRecord(
        provider=str(item.get("provider", "mouser")),
        distributor_part_number=str(item.get("distributor_part_number", "")),
        manufacturer_part_number=str(item.get("manufacturer_part_number", "")),
        manufacturer=str(item.get("manufacturer", "")),
        description=str(item.get("description", "")),
        availability=str(item.get("availability", "")),
        stock_quantity=str(item.get("stock_quantity", "")),
        price_breaks=price_breaks,
    )


def lookup_mouser_prices(rows: list[BomRow], mcp_url: str | None) -> dict[str, dict[str, object]]:
    resolved_mcp_url = (mcp_url or DEFAULT_MCP_URL).strip()

    if not asyncio.run(probe_mcp_mouser(resolved_mcp_url)):
        raise SystemExit(
            "Mouser is not available via MCP.\n"
            f"  Start the KiCad MCP server with MOUSER_API_KEY configured, then retry.\n"
            f"  MCP URL: {resolved_mcp_url}"
        )

    print(f"Using MCP server for Mouser lookups: {resolved_mcp_url}")
    return asyncio.run(lookup_mouser_prices_mcp(rows, resolved_mcp_url))


def style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table_title(ws, row: int, title: str) -> None:
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)


MOUSER_PLACED_HEADERS = [
    "Line",
    "References",
    "Value",
    "Manufacturer",
    "MPN",
    "BOM Qty / Board",
    "Order Qty (BOM × PCBA)",
    "Mouser Unit Price (USD)",
    "Line Total (USD)",
    "Mouser P/N",
    "Availability",
]


def write_mouser_placed_table(
    ws,
    *,
    current_row: int,
    title: str,
    items: list[tuple[BomRow, dict[str, object]]],
    pcba_cell: str,
    total_label: str,
) -> tuple[int, int | None]:
    """Write one Mouser placed-components table; return (next_row, total_row)."""
    if not items:
        return current_row, None

    write_table_title(ws, current_row, title)
    current_row += 1
    for col, header in enumerate(MOUSER_PLACED_HEADERS, start=1):
        style_header(ws.cell(row=current_row, column=col, value=header))
    current_row += 1
    first_data_row = current_row

    for row, info in items:
        breaks = usable_price_breaks(list(info.get("price_breaks") or []))
        ws.cell(row=current_row, column=1, value=row.line)
        ws.cell(row=current_row, column=2, value=row.references)
        ws.cell(row=current_row, column=3, value=row.value)
        ws.cell(row=current_row, column=4, value=row.manufacturer)
        ws.cell(row=current_row, column=5, value=row.mpn)
        ws.cell(row=current_row, column=6, value=row.quantity)
        order_qty_cell = ws.cell(row=current_row, column=7, value=f"=F{current_row}*{pcba_cell}")
        order_qty_cell.number_format = "0"
        order_qty_ref = f"G{current_row}"
        price_cell = ws.cell(
            row=current_row,
            column=8,
            value=excel_tiered_unit_price_formula(order_qty_ref, breaks),
        )
        price_cell.number_format = "$0.0000"
        total_cell = ws.cell(row=current_row, column=9, value=f"={order_qty_ref}*H{current_row}")
        total_cell.number_format = "$0.0000"
        ws.cell(row=current_row, column=10, value=str(info.get("mouser_pn", "")))
        ws.cell(row=current_row, column=11, value=format_availability_label(info))
        current_row += 1

    last_data_row = current_row - 1
    ws.cell(row=current_row, column=5, value=total_label).font = Font(bold=True)
    total_cell = ws.cell(
        row=current_row,
        column=9,
        value=f"=SUM(I{first_data_row}:I{last_data_row})",
    )
    total_cell.font = Font(bold=True)
    total_cell.number_format = "$0.00"
    total_row = current_row
    return current_row + 2, total_row


def build_workbook(
    bom_rows: list[BomRow],
    lookup: dict[str, dict[str, object]],
    output_path: Path,
    source_name: str,
) -> None:
    placed = [row for row in bom_rows if not row.is_dnp]
    dnp_rows = [row for row in bom_rows if row.is_dnp]

    instock: list[tuple[BomRow, dict[str, object]]] = []
    outstock: list[tuple[BomRow, dict[str, object]]] = []
    unavailable: list[BomRow] = []
    for row in placed:
        info = lookup.get(row.mpn, empty_lookup_entry())
        if info.get("found"):
            if is_mouser_in_stock(info):
                instock.append((row, info))
            else:
                outstock.append((row, info))
        else:
            unavailable.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Cost"

    ws["A1"] = "BOM Cost Estimate (Mouser)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Source BOM"
    ws["B2"] = source_name
    ws["A3"] = "PCBA Quantity"
    ws["B3"] = 1
    ws["B3"].number_format = "0"
    ws["A3"].font = Font(bold=True)
    ws["B3"].fill = PatternFill("solid", fgColor="FFF2CC")

    pcba_cell = "B3"
    current_row = 5

    current_row, mouser_instock_total_row = write_mouser_placed_table(
        ws,
        current_row=current_row,
        title="Placed Components — Available on Mouser (In Stock)",
        items=instock,
        pcba_cell=pcba_cell,
        total_label="Total Mouser (instock)",
    )
    current_row, mouser_outstock_total_row = write_mouser_placed_table(
        ws,
        current_row=current_row,
        title="Placed Components — Available on Mouser (Out of Stock)",
        items=outstock,
        pcba_cell=pcba_cell,
        total_label="Total Mouser (outstock)",
    )
    if not instock and not outstock:
        current_row += 1

    write_table_title(ws, current_row, "Placed Components — Not Available on Mouser (Unit Price = 0)")
    current_row += 1
    unavailable_headers = [
        "Line",
        "References",
        "Value",
        "Manufacturer",
        "MPN",
        "BOM Qty",
        "Unit Price (USD)",
        "Line Total / Board (USD)",
    ]
    for col, title in enumerate(unavailable_headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=title)
        style_header(cell)
    current_row += 1
    unavailable_first_data_row = current_row
    for row in unavailable:
        ws.cell(row=current_row, column=1, value=row.line)
        ws.cell(row=current_row, column=2, value=row.references)
        ws.cell(row=current_row, column=3, value=row.value)
        ws.cell(row=current_row, column=4, value=row.manufacturer)
        ws.cell(row=current_row, column=5, value=row.mpn)
        ws.cell(row=current_row, column=6, value=row.quantity)
        zero_cell = ws.cell(row=current_row, column=7, value=0)
        zero_cell.number_format = "$0.0000"
        line_total = ws.cell(row=current_row, column=8, value=f"=F{current_row}*G{current_row}")
        line_total.number_format = "$0.0000"
        current_row += 1

    unavailable_last_data_row = current_row - 1
    non_mouser_total_row: int | None = None
    if unavailable:
        ws.cell(row=current_row, column=5, value="Total (Not on Mouser × PCBA Qty)").font = Font(bold=True)
        non_mouser_total_cell = ws.cell(
            row=current_row,
            column=8,
            value=f"=SUM(H{unavailable_first_data_row}:H{unavailable_last_data_row})*{pcba_cell}",
        )
        non_mouser_total_cell.font = Font(bold=True)
        non_mouser_total_cell.number_format = "$0.00"
        non_mouser_total_row = current_row
        current_row += 2
    else:
        current_row += 1

    write_table_title(ws, current_row, "Combined BOM Cost Summary")
    current_row += 1
    mouser_instock_total_ref = f"I{mouser_instock_total_row}" if mouser_instock_total_row else "0"
    mouser_outstock_total_ref = f"I{mouser_outstock_total_row}" if mouser_outstock_total_row else "0"
    non_mouser_total_ref = f"H{non_mouser_total_row}" if non_mouser_total_row else "0"

    ws.cell(row=current_row, column=5, value="Total — Mouser (instock)").font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=f"={mouser_instock_total_ref}").number_format = "$0.00"
    mouser_instock_summary_row = current_row
    current_row += 1

    ws.cell(row=current_row, column=5, value="Total — Mouser (outstock)").font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=f"={mouser_outstock_total_ref}").number_format = "$0.00"
    mouser_outstock_summary_row = current_row
    current_row += 1

    ws.cell(row=current_row, column=5, value="Total — Not on Mouser (× PCBA Qty)").font = Font(bold=True)
    ws.cell(row=current_row, column=9, value=f"={non_mouser_total_ref}").number_format = "$0.00"
    non_mouser_summary_row = current_row
    current_row += 1

    ws.cell(row=current_row, column=5, value="Combined Total (All PCBA)").font = Font(bold=True)
    combined_total_cell = ws.cell(
        row=current_row,
        column=9,
        value=(
            f"=I{mouser_instock_summary_row}+I{mouser_outstock_summary_row}+I{non_mouser_summary_row}"
        ),
    )
    combined_total_cell.font = Font(bold=True, size=12)
    combined_total_cell.number_format = "$0.00"
    combined_total_cell.fill = PatternFill("solid", fgColor="DDEBF7")
    combined_total_row = current_row
    current_row += 1

    ws.cell(row=current_row, column=5, value="Grand Total").font = Font(bold=True)
    grand_total_cell = ws.cell(
        row=current_row,
        column=9,
        value=f"=I{combined_total_row}",
    )
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.number_format = "$0.00"
    grand_total_cell.fill = PatternFill("solid", fgColor="E2EFDA")
    current_row += 3

    write_table_title(ws, current_row, "Do Not Place (DNP)")
    current_row += 1
    dnp_headers = ["Line", "References", "Value", "Manufacturer", "MPN", "BOM Qty", "Note"]
    for col, title in enumerate(dnp_headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=title)
        style_header(cell)
    current_row += 1
    for row in dnp_rows:
        ws.cell(row=current_row, column=1, value=row.line)
        ws.cell(row=current_row, column=2, value=row.references)
        ws.cell(row=current_row, column=3, value=row.value)
        ws.cell(row=current_row, column=4, value=row.manufacturer)
        ws.cell(row=current_row, column=5, value=row.mpn)
        ws.cell(row=current_row, column=6, value=row.quantity)
        ws.cell(row=current_row, column=7, value=row.note or "Do not place")
        current_row += 1

    widths = [8, 28, 18, 22, 28, 12, 16, 18, 16, 18, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A6"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_updated{output_path.suffix}")
        wb.save(fallback)
        print(f"Could not overwrite locked file. Saved to {fallback}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build BOM cost Excel from KiCad BOM CSV.")
    parser.add_argument("bom_csv", type=Path, help="Path to KiCad BOM CSV export")
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        type=Path,
        help="Output Excel path (default: <bom>_bom_cost_mouser.xlsx beside CSV)",
    )
    parser.add_argument(
        "--mcp-url",
        default=DEFAULT_MCP_URL,
        help=f"KiCad MCP server URL (default: {DEFAULT_MCP_URL})",
    )
    args = parser.parse_args()

    bom_path = args.bom_csv.expanduser().resolve()
    if not bom_path.is_file():
        print(f"BOM file not found: {bom_path}", file=sys.stderr)
        return 1

    output_path = (
        args.output_xlsx.expanduser().resolve()
        if args.output_xlsx
        else bom_path.with_name(f"{bom_path.stem}_bom_cost_mouser.xlsx")
    )

    bom_rows = parse_bom_csv(bom_path)
    placed_rows = [row for row in bom_rows if not row.is_dnp]
    print(f"Parsed {len(bom_rows)} BOM lines ({len(placed_rows)} placed, {len(bom_rows) - len(placed_rows)} DNP)")

    lookup = lookup_mouser_prices(bom_rows, args.mcp_url.strip())
    build_workbook(bom_rows, lookup, output_path, bom_path.name)

    available_count = sum(1 for row in placed_rows if lookup.get(row.mpn, {}).get("found"))
    print(f"Wrote {output_path}")
    print(f"Mouser matches: {available_count}/{len(placed_rows)} placed lines")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
